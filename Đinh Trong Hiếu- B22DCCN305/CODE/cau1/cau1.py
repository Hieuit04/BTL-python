# Import các thư viện cần thiết
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from pathlib import Path
# Phần 1: Thu thập dữ liệu chính
def collect_main_data():
    url = "https://fbref.com/en/comps/9/2023-2024/stats/2023-2024-Premier-League-Stats"
    
    # Khởi tạo trình duyệt
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver.get(url)
    driver.implicitly_wait(10)
    
    # Lấy nội dung trang web
    page_content = driver.page_source
    driver.quit()
    
    # Phân tích nội dung trang web
    soup = BeautifulSoup(page_content, 'html.parser')
    table = soup.find('table', {'id': 'stats_standard'})
    
    if table:
        players_data = []
        
        for row in table.find_all('tr', {'data-row': True}):
            player_data = {}
            for cell in row.find_all(['th', 'td']):
                stat = cell.get('data-stat')
                value = cell.text.strip()
                csk_value = cell.get('csk')  # Lấy giá trị csk
                
                if stat == 'minutes':
                    player_data['csk_minutes'] = int(csk_value.replace(',', '')) if csk_value else 0
                player_data[stat] = value

            players_data.append(player_data)

        # Chuyển dữ liệu thành DataFrame
        players_df = pd.DataFrame(players_data)
        players_df['Age'] = pd.to_numeric(players_df['age'], errors='coerce')
        filtered_players_df = players_df[players_df['csk_minutes'] > 90]
        
        # Tách tên thành first name và last name
        filtered_players_df['first_name'] = filtered_players_df['player'].apply(lambda x: x.split()[0])
        filtered_players_df['last_name'] = filtered_players_df['player'].apply(lambda x: x.split()[-1])
        
        filtered_players_df = filtered_players_df.sort_values(by=['last_name', 'Age'], ascending=[True, True])
        filtered_players_df = filtered_players_df.drop(columns=['Age', 'first_name', 'last_name'])
        
        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(filtered_players_df, index=False, header=True):
            ws.append(r)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        wb.save('results.xlsx')
        print("Dữ liệu đã được xuất ra file 'results.xlsx'")
        
        return set(filtered_players_df['player'])
    else:
        print("Không tìm thấy bảng dữ liệu")
        return set()

# Phần 2: Thu thập dữ liệu bổ sung
def collect_additional_data(player_set):
    url_table_mapping = [
        ("https://fbref.com/en/comps/9/2023-2024/keepers/2023-2024-Premier-League-Stats", "stats_keeper"),
        ("https://fbref.com/en/comps/9/2023-2024/defense/2023-2024-Premier-League-Stats", "stats_defense"),
        ("https://fbref.com/en/comps/9/2023-2024/gca/2023-2024-Premier-League-Stats", "stats_gca"),
        ("https://fbref.com/en/comps/9/2023-2024/misc/2023-2024-Premier-League-Stats", "stats_misc"),
        ("https://fbref.com/en/comps/9/2023-2024/passing_types/2023-2024-Premier-League-Stats", "stats_passing_types"),
        ("https://fbref.com/en/comps/9/2023-2024/passing/2023-2024-Premier-League-Stats", "stats_passing"),
        ("https://fbref.com/en/comps/9/2023-2024/playingtime/2023-2024-Premier-League-Stats", "stats_playing_time"),
        ("https://fbref.com/en/comps/9/2023-2024/possession/2023-2024-Premier-League-Stats", "stats_possession"),
        ("https://fbref.com/en/comps/9/2023-2024/shooting/2023-2024-Premier-League-Stats", "stats_shooting"),
    ]

    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

    for url, table_id in url_table_mapping:
        driver.get(url)
        driver.implicitly_wait(10)
        page_content = driver.page_source
        soup = BeautifulSoup(page_content, 'html.parser')
        table = soup.find('table', {'id': table_id})

        if table:
            players_data = []

            for row in table.find_all('tr', {'data-row': True}):
                player_data = {}
                for cell in row.find_all(['th', 'td']):
                    stat = cell.get('data-stat')
                    value = cell.text.strip()
                    player_data[stat] = value

                players_data.append(player_data)

            players_df = pd.DataFrame(players_data)
            filtered_players_df = players_df[players_df['player'].isin(player_set)]
            
            wb = Workbook()
            ws = wb.active

            header = [col for col in filtered_players_df.columns if col != 'minutes']
            ws.append(header)

            for row_data in filtered_players_df.values.tolist():
                row_without_minutes = [value for col_name, value in zip(filtered_players_df.columns, row_data) if col_name != 'minutes']
                ws.append(row_without_minutes)

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            wb.save(f'{table_id}.xlsx')
            print(f"Dữ liệu đã được xuất ra file '{table_id}.xlsx'")
        else:
            print(f"Không tìm thấy bảng dữ liệu với ID '{table_id}'")

    driver.quit()

# Phần 3: Hợp nhất các file Excel
def merge_data_files():
    results_df = pd.read_excel('results.xlsx')
    keeper_df = pd.read_excel('stats_keeper.xlsx')
    passing_types_df = pd.read_excel('stats_passing_types.xlsx')
    possession_df = pd.read_excel('stats_possession.xlsx')
    gca_df = pd.read_excel('stats_gca.xlsx')
    defense_df = pd.read_excel('stats_defense.xlsx')
    misc_df = pd.read_excel('stats_misc.xlsx')
    passing_df = pd.read_excel('stats_passing.xlsx')
    playing_time_df = pd.read_excel('stats_playing_time.xlsx')
    shooting_df = pd.read_excel('stats_shooting.xlsx')

    merged_df = results_df.merge(keeper_df, on=['player', 'team'], how='outer', suffixes=('', '_dup'))
    merged_df = merged_df.merge(passing_types_df, on=['player', 'team'], how='outer', suffixes=('', '_dup'))
    merged_df = merged_df.merge(possession_df, on=['player', 'team'], how='outer', suffixes=('', '_dup'))
    merged_df = merged_df.merge(gca_df, on=['player', 'team'], how='outer', suffixes=('', '_dup'))
    merged_df = merged_df.merge(defense_df, on=['player', 'team'], how='outer', suffixes=('', '_dup'))
    merged_df = merged_df.merge(misc_df, on=['player', 'team'], how='outer', suffixes=('', '_dup'))
    merged_df = merged_df.merge(passing_df, on=['player', 'team'], how='outer', suffixes=('', '_dup'))
    merged_df = merged_df.merge(playing_time_df, on=['player', 'team'], how='outer', suffixes=('', '_dup'))
    merged_df = merged_df.merge(shooting_df, on=['player', 'team'], how='outer', suffixes=('', '_dup'))

    merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]
    merged_df['last_name'] = merged_df['player'].apply(lambda x: x.split()[-1] if isinstance(x, str) else '')
    merged_df = merged_df.sort_values(by=['last_name', 'csk_minutes'])
    merged_df = merged_df.drop(columns=['last_name'])
    merged_df = merged_df[merged_df['ranker'].notnull()]
    
    merged_df.to_excel("merged_results.xlsx", index=False)
    print("Dữ liệu đã được hợp nhất và lưu vào file 'merged_results.xlsx'")

# Chạy các phần
if __name__ == "__main__":
    player_set = collect_main_data()
    if player_set:
        collect_additional_data(player_set)
        merge_data_files()
# Đường dẫn file đầu vào và đầu ra
input_path = 'merged_results.xlsx'
output_path = 'results.xlsx'

# Đọc dữ liệu từ file Excel
df = pd.read_excel(input_path)

# Mapping tên cột cũ sang tên cột mới
column_mapping = {
    #Standard Stats
    #-" "
    'ranker':'Rk',
    'player':'Player',
    'nationality': 'Nation',
    'team': 'Team',
    'position': 'Position',
    'age': 'Age',
    'games': 'Matches played',
    'games_starts': 'Starts',
    'minutes': 'Minutes',
    #-performance
    'goals_pens': 'Non-Penalty Goals',
    'pens_made': 'Penalty Goals',
    'assists': 'Assists',
    'cards_yellow': 'Yellow Cards',
    'cards_red': 'Red Cards',
    #-Expected
    'xg': 'xG',
    'npxg': 'npxG',
    'xg_assist': 'xAG',
    #-Progression
    'progressive_carries': 'PrgC',
    'progressive_passes': 'PrgP',
    'progressive_passes_received': 'PrgR',
    #-Per 90 minutes
    'goals_per90': 'Gls/90',
    'assists_per90': 'Ast/90',
    'goals_assists_per90': 'G+A/90',
    'goals_pens_per90': 'G-PK/90',
    'goals_assists_pens_per90': 'G+A-PK/90',
    'xg_per90': 'xG/90',
    'xg_assist_per90': 'xAG/90',
    'xg_xg_assist_per90': 'xG + xAG/90',
    'npxg_per90': 'npxG/90',
    'npxg_xg_assist_per90': 'npxG + xAG/90',
    #Goalkeeping
    #-Performance:
    'gk_goals_against': 'GA',
    'gk_goals_against_per90': 'GA90',
    'gk_shots_on_target_against': 'SoTA',
    'gk_saves': 'Saves',
    'gk_save_pct': 'Save%-Performance',
    'gk_wins': 'W',
    'gk_ties': 'D',
    'gk_losses': 'L',
    'gk_clean_sheets': 'CS',
    'gk_clean_sheets_pct': 'CS%',
    #-Penalty Kicks
    'gk_pens_att': 'PKatt-PK',
    'gk_pens_allowed': 'PKA',
    'gk_pens_saved': 'PKsv',
    'gk_pens_missed': 'PKm',
    'gk_pens_save_pct': 'Save%/PK',

    # Shooting:
    #-Standard:
    'goals_dup': 'Gls',  
    'shots': 'Sh',
    'shots_on_target': 'SoT',
    'shots_on_target_pct': 'SoT%',
    'shots_per90': 'Sh/90',
    'shots_on_target_per90': 'SoT/90',
    'goals_per_shot': 'G/Sh',
    'goals_per_shot_on_target': 'G/SoT',
    'average_shot_distance': 'Dist',
    'shots_free_kicks': 'FK-Shooting',
    'pens_made_dup': 'PK',
    'pens_att_dup': 'PKatt-Standard',
    #-Expected
    'xg_dup': 'xG-Shooting',
    'npxg_dup': 'npxG-Shooting',
    'npxg_per_shot': 'npxG/Sh',
    'xg_net': 'G-xG',
    'npxg_net': 'np:xG-xG',
    #Passing
    #-Total
    'passes_completed': 'Cmp/Total',
    'passes': 'Att/Total',
    'passes_pct': 'Cmp%/Total',
    'passes_total_distance': 'TotDist/Total',
    'passes_progressive_distance': 'PrgDist/Total',
    #-Short
    'passes_completed_short': 'Cmp/Short',
    'passes_short': 'Att/Short',
    'passes_pct_short': 'Cmp%/Short',
    #-Medium
    'passes_completed_medium': 'Cmp/Medium',
    'passes_medium': 'Att/Medium',
    'passes_pct_medium': 'Cmp%/Medium',
    #-Long
    'passes_completed_long': 'Cmp/Long',
    'passes_long': 'Att/Long',
    'passes_pct_long': 'Cmp%/Long',
    #-Expected
    'assists_dup': 'Ast',
    'xg_assist_dup': 'xAG-pass',
    'pass_xa': 'xA',
    'xg_assist_net': 'A-xAG',
    'assisted_shots': 'KP',
    'passes_into_final_third': '1/3-pass',
    'passes_into_penalty_area': 'PPA',
    'crosses_into_penalty_area': 'CrsPA',
    'progressive_passes_dup': 'PrgP-pass',
    #Pass Types:
    #Pass Types:
    'passes_live': 'Live',
    'passes_dead': 'Dead',
    'passes_free_kicks': 'FK',
    'through_balls': 'TB',
    'passes_switches': 'Sw',
    'crosses': 'Crs-PT',
    'throw_ins': 'TI',
    'corner_kicks': 'CK',
    #-Corner Kicks
    'corner_kicks_in': 'In',
    'corner_kicks_out': 'Out',
    'corner_kicks_straight': 'Str',
    #-Outcomes
    'passes_completed': 'Cmp-OC',
    'passes_offsides': 'Off-OC',
    'passes_blocked': 'Blocks-OC',

    #Goal and Shot Creation:
    #-SCA
    'sca': 'SCA',
    'sca_per90': 'SCA90',
    #-SCA Types
    'sca_passes_live': 'PassLive/SCA',
    'sca_passes_dead': 'PassDead/SCA',
    'sca_take_ons': 'TO/SCA',
    'sca_shots': 'Sh/SCA',
    'sca_fouled': 'Fld/SCA',
    'sca_defense': 'Def/SCA',
    #-GCA
    'gca': 'GCA',
    'gca_per90': 'GCA90',
    #-GCA Types
    'gca_passes_live': 'PassLive/GCA',
    'gca_passes_dead': 'PassDead/GCA',
    'gca_take_ons': 'TO/GCA',
    'gca_shots': 'Sh/GCA',
    'gca_fouled': 'Fld/GCA',
    'gca_defense': 'Def/GCA',
    #tackles
    #-Tackles
    'tackles': 'Tkl-Tackles',
    'tackles_won': 'TklW',
    'tackles_def_3rd': 'Def 3rd-Tackles',
    'tackles_mid_3rd': 'Mid 3rd-Tackles',
    'tackles_att_3rd': 'Att 3rd-Tackles',
    #-Challenges
    'challenge_tackles': 'Tkl-Challenges',
    'challenges': 'Att-Challenges',
    'challenge_tackles_pct': 'Tkl%-Challenges',
    'challenges_lost': 'Lost-Challenges',
    #-Blocks
    'blocks': 'Blocks',
    'blocked_shots': 'Sh-Blocks',
    'blocked_passes': 'Pass',
    'interceptions': 'Int',
    'tackles_interceptions': 'Tkl+Int',
    'clearances': 'Clr',
    'errors': 'Err',
    #Possession
    #-Touches
    'touches': 'Touches',
    'touches_def_pen_area': 'Def Pen',
    'touches_def_3rd': 'Def 3rd-Touches',
    'touches_mid_3rd': 'Mid 3rd-Touches',
    'touches_att_3rd': 'Att 3rd-Touches',
    'touches_att_pen_area': 'Att Pen',
    'touches_live_ball': 'Live-Touches',
    #-Take-Ons
    'take_ons': 'Att-Take-Ons',
    'take_ons_won': 'Succ',
    'take_ons_won_pct': 'Succ%',
    'take_ons_tackled': 'Tkld',
    'take_ons_tackled_pct': 'Tkld%',
    #-Carries
    'carries': 'Carries',
    'carries_distance': 'TotDist-Carries',
    'carries_progressive_distance': 'ProDist',
    'progressive_carries_dup': 'ProgC',
    'carries_into_final_third': '1/3-Carries',
    'carries_into_penalty_area': 'CPA',
    'miscontrols': 'Mis',
    'dispossessed': 'Dis',
    #-Receiving
    'passes_received': 'Rec',
    'progressive_passes_received_dup': 'PrgR-Receiving',
    #Playing Time
    #-Starts
    'games_starts_dup': 'Starts/PT',
    'minutes_per_start': 'Mn/Start',
    'games_complete': 'Compl',
    #-Subs
    'games_subs': 'Subs',
    'minutes_per_sub': 'Mn/Sub',
    'unused_subs': 'Unused Subs',
    #-Team Success
    'points_per_game': 'PPM',
    'on_goals_for': 'onG',
    'on_goals_against': 'onGA',
    #-Team Success xG
    'on_xg_for': 'onxG',
    'on_xg_against': 'onxGA',
    #Miscellaneous Stats:
    #-Performance
    'fouls': 'Fls',
    'fouled': 'Fld',
    'offsides': 'Off/PE',
    'crosses_dup': 'Crs/PE',
    'own_goals': 'OG',
    'ball_recoveries': 'Recov',
    #-Aerial Duels
    'aerials_won': 'Won',
    'aerials_lost': 'Lost/AD',
    'aerials_won_pct': 'Won%',
}

# Đổi tên các cột
df.rename(columns=column_mapping, inplace=True)

# Nhóm các cột theo yêu cầu
# Nhóm các cột theo yêu cầu
group_columns = {
    ' ':{
    ' ': ['Rk', 'Player', 'Nation', 'Team', 'Position', 'Age','Matches played','Starts','Minutes'],
    'Performance ': ['Non-Penalty Goals','Penalty Goals','Assists','Yellow Cards','Red Cards'],
    'Expected ': ['xG','npxG','xAG',],
    'Progression': ['PrgC','PrgP','PrgR'],
    'Per 90 Minutes': ['Gls/90', 'Ast/90', 'G+A/90', 'G-PK/90', 'G+A-PK/90', 'xG/90', 'xAG/90', 'xG + xAG/90', 'npxG/90', 'npxG + xAG/90']
    },
    'Goalkeeping':{
    'Performance/GK': ['GA', 'GA90', 'SoTA', 'Saves', 'Save%-Performance', 'W', 'D', 'L', 'CS', 'CS%'],
    'Penalty Kicks/GK': ['PKatt-PK', 'PKA', 'PKsv', 'PKm', 'Save%/PK']
    },
    
    'Shooting':{
    'Standard/SH': ['Gls', 'Sh', 'SoT', 'SoT%', 'Sh/90', 'SoT/90', 'G/Sh', 'G/SoT', 'Dist', 'FK-Shooting', 'PK', 'PKatt-Standard'],
    'Expected/SH': ['xG-Shooting', 'npxG-Shooting', 'npxG/Sh', 'G-xG','np:xG-xG']
    },
    'Passing':{
    'Total/PASS': ['Cmp/Total', 'Att/Total', 'Cmp%/Total', 'TotDist/Total', 'PrgDist/Total',],
    'Short/PASS':['Cmp/Short', 'Att/Short', 'Cmp%/Short',],
    'Medium/PASS':['Cmp/Medium', 'Att/Medium', 'Cmp%/Medium',],
    'Long/PASS':['Cmp/Long', 'Att/Long', 'Cmp%/Long'],
    'Expected/PASS': ['Ast', 'xAGpass', 'xA', 'A-xAG', 'KP', '1/3-pass', 'PPA', 'CrsPA', 'PrgP-pass']
    },
    'Pass Types':{
    'Pass Types/PT:': ['Live', 'Dead', 'FK/PT', 'TB', 'Sw','Crs-PT','TI','CK',],
    'Corner Kicks/PT': ['In','Out','Str'],
    'Outcomes/PT': ['Cmp-OC','Off-OC','Blocks-OC']
    },
    'Goal and Shot Creation':{
    'SCA/GASC': ['SCA','SCA90'],
    'SCA Types/GASC': ['PassLive/SCA','PassDead/SCA','TO/SCA','Sh/SCA','Fld/SCA','Def/SCA',],
    'GCA/GASC': ['GCA','GCA90',],
    'GCA Types/GASC': ['PassLive/GCA','PassDead/GCA','TO/GCA','Sh/GCA','Fld/GCA','Def/GCA',]
    },
    'Tackles':{
    'Tackles': ['Tkl/Tackles','TklW','Def 3rd-Tackles','Mid 3rd-Tackles','Att 3rd-Tackles'],
    'Challenges': ['Tkl-Challenges','Att-Challenges','Tkl%-Challenges','Lost-Challenges'],
    'Blocks': ['Blocks','Sh-Blocks','Pass','Int','Tkl+Int','Clr','Err']
    },
    'Possession':{
    'Touches': ['Touches','Def Pen','Def 3rd-Touches','Mid 3rd-Touches','Att 3rd-Touches','Att Pen','Live-Touches'],
    'Take-Ons': ['Att-Take-Ons','Succ','Succ%','Tkld','Tkld%'],
    'Carries': ['Carries','TotDist-Carries','ProDist','ProgC','1/3-Carries','CPA','Mis','Dis'],
    'Receiving': ['Rec','PrgR-Receiving']
    },
    'Playing Time':{
    'Starts': ['Starts/PT','Mn/Start','Compl'],
    'Subs': ['Subs','Mn/Sub','Unused Subs'],
    'Team Success': ['PPM','onG','onGA'],
    'Team Success xG': ['onxG','onxGA']
    },
    'Miscellaneous Stats':{
    'Performance/MS': ['Fls','Fld','Off/PE','Crs/PE','OG','Recov'],
    'Aerial Duels': ['Won','Lost/AD','Won%']
    }   
}


# Tạo DataFrame mới với MultiIndex
new_df = pd.DataFrame()

# Thêm dữ liệu vào DataFrame mới với MultiIndex
for group_name, subgroups in group_columns.items():
    for subgroup_name, columns in subgroups.items():
        if set(columns).issubset(df.columns):
            grouped_data = df[columns]
            # Tạo MultiIndex với nhóm lớn và nhóm nhỏ
            grouped_data.columns = pd.MultiIndex.from_product([[group_name], [subgroup_name], grouped_data.columns])
            new_df = pd.concat([new_df, grouped_data], axis=1)

# Lưu vào file Excel
new_df.to_excel(output_path, index=True)

print("File đã được lưu thành công:", output_path)

# Tìm và xóa các file Excel có đuôi .xlsx
for file in Path(".").glob("*.xlsx"):
    if file.name != "results.xlsx":  # giữ lại file 'results.xlsx'
        file.unlink()
print("Đã xóa tất cả các file Excel trừ file 'results.xlsx'")

# Đường dẫn đến file Excel
input_file_path = 'results.xlsx'  # Thay đổi đường dẫn cho phù hợp
output_file_path = 'results.csv'               # Đường dẫn file CSV đầu ra

# Đọc file Excel
df = pd.read_excel(input_file_path)

# Xuất dữ liệu ra file CSV
df.to_csv(output_file_path, index=False)  # index=False để không ghi chỉ số dòng vào file CSV

print(f"Đã chuyển đổi thành công {input_file_path} sang {output_file_path}")